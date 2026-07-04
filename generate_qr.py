import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
import requests
import json
import urllib.parse
import io
import os
import time
import qrcode
from PIL import Image, ImageDraw

# Read groupQuestions.js to get our group IDs
with open('src/data/groupQuestions.js', 'r', encoding='utf-8') as f:
    content = f.read()

our_ids = sorted(set(re.findall(r'"groupId": "([^"]+)"', content)))

# Read Canva Code.xlsx
wb_canva = openpyxl.load_workbook('C:/Users/s20201020/Downloads/Canva Code.xlsx')
ws_canva = wb_canva.active

canva_ids = []
for row in ws_canva.iter_rows(min_row=1, max_row=ws_canva.max_row, min_col=1, max_col=1, values_only=True):
    if row[0]:
        canva_ids.append(str(row[0]).strip())

# Build mapping from Canva ID to our group ID
def canva_to_our(canva_id):
    if canva_id.startswith('G'):
        m = re.match(r'(G\d+)([A-Z])(\d+)$', canva_id)
        if m:
            grade, cls, num = m.groups()
            result = f"{grade}{cls}-{int(num)}"
            if result in our_ids:
                return result
        m = re.match(r'(G\d+)([A-Z])(\d)(\d+)$', canva_id)
        if m:
            grade, cls, class_num, group_num = m.groups()
            result = f"{grade}{cls}{class_num}-{int(group_num)}"
            if result in our_ids:
                return result
    else:
        if canva_id.startswith('SIN'):
            result = f"Singapore-{int(canva_id[3:])}"
            if result in our_ids: return result
        elif canva_id.startswith('SC'):
            result = f"Sichuan-{int(canva_id[2:])}"
            if result in our_ids: return result
        elif canva_id.startswith('KL'):
            result = f"Malaysia-{int(canva_id[2:])}"
            if result in our_ids: return result
        elif canva_id.startswith('SP'):
            result = f"Madrid-{int(canva_id[2:])}"
            if result in our_ids: return result
        elif canva_id.startswith('LON'):
            result = f"London-{int(canva_id[3:])}"
            if result in our_ids: return result
        elif canva_id.startswith('SH'):
            result = f"Shanghai-{int(canva_id[2:])}"
            if result in our_ids: return result
    return None

# Group by grade
grade_canva = {}
for cid in canva_ids:
    mapped = canva_to_our(cid)
    if mapped:
        if cid.startswith('G'):
            grade = cid[:2]
        else:
            grade = 'E'
        if grade not in grade_canva:
            grade_canva[grade] = []
        grade_canva[grade].append((cid, mapped))

# QRCode Monkey API config
QR_CONFIG = {
    'body': 'circle',
    'eye': 'frame13',
    'eyeBall': 'ball14',
    'bodyColor': '#009E20',
    'bg': '#ffffff'
}

def generate_qr_monkey(group_id):
    """Generate QR code using QRCode Monkey API"""
    data = f"https://hkbuaspbl2026.pages.dev/?id={group_id}"
    url = f'https://api.qrcode-monkey.com/qr/custom?data={urllib.parse.quote(data)}&config={urllib.parse.quote(json.dumps(QR_CONFIG))}&size=300&file=png&download=false'
    
    try:
        response = requests.get(url, timeout=30)
        if response.status_code == 200 and response.headers.get('content-type', '').startswith('image'):
            return response.content
        return None
    except Exception as e:
        return None

def generate_qr_local(group_id):
    """Generate QR code locally with green circular data modules and square finder patterns"""
    qr = qrcode.QRCode(
        version=3,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=12,
        border=2,
    )
    qr.add_data(f"https://hkbuaspbl2026.pages.dev/?id={group_id}")
    qr.make(fit=True)
    
    matrix = qr.get_matrix()
    box_size = 12
    border = 2
    matrix_size = len(matrix)
    img_size = matrix_size * box_size + 2 * border * box_size
    img = Image.new('RGB', (img_size, img_size), (255, 255, 255))
    draw = ImageDraw.Draw(img)
    
    GREEN = (0, 158, 32)
    
    for row_idx, row in enumerate(matrix):
        for col_idx, cell in enumerate(row):
            if cell:
                x = col_idx * box_size + border * box_size
                y = row_idx * box_size + border * box_size
                
                size = len(matrix)
                is_finder = False
                
                # Check if this is part of a finder pattern (the 7x7 corner squares)
                if row_idx < 7 and col_idx < 7:
                    is_finder = True
                elif row_idx < 7 and col_idx >= size - 7:
                    is_finder = True
                elif row_idx >= size - 7 and col_idx < 7:
                    is_finder = True
                
                if is_finder:
                    # Finder patterns MUST be square for scanners to work
                    draw.rectangle([x, y, x+box_size-1, y+box_size-1], fill=GREEN)
                else:
                    # Data modules can be circular
                    draw.ellipse([x+1, y+1, x+box_size-2, y+box_size-2], fill=GREEN)
    
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    return buf.getvalue()

# Grade names
grade_names = {
    'G1': 'Grade_1',
    'G2': 'Grade_2',
    'G3': 'Grade_3',
    'G4': 'Grade_4',
    'G5': 'Grade_5',
    'G6': 'Grade_6',
    'E': 'Excursion'
}

# Style definitions
header_fill = PatternFill(start_color='009E20', end_color='009E20', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=12)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
center_align = Alignment(horizontal='center', vertical='center')

# Create QR code images folder
os.makedirs('QR codes/images', exist_ok=True)

# Create Excel files with embedded QR codes
for grade, entries in sorted(grade_canva.items()):
    grade_name = grade_names[grade]
    filename = f'QR codes_new/{grade_name}.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = grade_name
    
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    
    headers = ['Canva Group ID', 'QR Code Image']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    ws.row_dimensions[1].height = 25
    
    for row, (canva_id, our_id) in enumerate(entries, 2):
        qr_url = f"https://hkbuaspbl2026.pages.dev/?id={our_id}"
        
        cell = ws.cell(row=row, column=1, value=canva_id)
        cell.border = border
        cell.alignment = center_align
        
        print(f"Generating QR for {canva_id} -> {our_id}...")
        
        # Try Monkey API first, fallback to local
        qr_bytes = generate_qr_monkey(our_id)
        if qr_bytes:
            print(f"  [Monkey API OK]")
        else:
            qr_bytes = generate_qr_local(our_id)
            print(f"  [Local fallback - circular]")
        
        if qr_bytes:
            img_path = f'QR codes/images/{canva_id}.png'
            with open(img_path, 'wb') as f:
                f.write(qr_bytes)
            
            img = XLImage(io.BytesIO(qr_bytes))
            img.width = 120
            img.height = 120
            ws.add_image(img, f'B{row}')
        else:
            ws.cell(row=row, column=2, value="Failed to generate")
        
        ws.row_dimensions[row].height = 100
        time.sleep(0.2)
    
    ws.freeze_panes = 'A2'
    wb.save(filename)
    print(f"Saved: {filename} ({len(entries)} entries)\n")

print("All Excel files created successfully!")
